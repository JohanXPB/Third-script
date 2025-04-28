import streamlit as st
import pandas as pd
import os
from io import StringIO
import sys

# Check if openpyxl is installed, if not, install it
try:
    import openpyxl
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

# Get the current working directory
current_directory = os.getcwd()

# Get all .dat files in the current directory
dat_files = [f for f in os.listdir(current_directory) if f.endswith(".dat")]

for file_name in dat_files:
    file_path = os.path.join(current_directory, file_name)

    # Read the dat file into a pandas DataFrame starting from the line following "Calibration"
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
            start_line = 0
            meta_data = []
            for i, line in enumerate(lines):
                if "Calibration: Calibration" in line:
                    start_line = i - 1  # Start reading from the line following "Calibration"
                    meta_data = lines[:start_line]
                    print(f"Start line: {start_line}")  # Print the start_line increments
                    break
        df = pd.read_csv(file_path, sep='\t', skiprows=start_line)

        # Redirect print output to a buffer
        buffer = StringIO()
        sys.stdout = buffer

        # Use the original file_path as the naming for the excel file
        file_name = os.path.splitext(os.path.basename(file_path))[0]

        # Convert the DataFrame to an excel file without writing the header
        excel_save_path = os.path.join(current_directory, file_name + "_analyzed_data.xlsx")
        df.to_excel(excel_save_path, index=False, header=False)
        print(f"The Excel file for {file_name} has been saved successfully.")

        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_save_path)
        sheet = workbook.active

        # Calculate the average and write the result to column K
        processed_strings = []
        for row in range(2, sheet.max_row + 1):
            current_value = str(sheet.cell(row, 1).value).split('_')[0]
            next_value = str(sheet.cell(row + 1, 1).value).split('_')[0] if row < sheet.max_row else ""
            if current_value != next_value:
                processed_strings.append(current_value)
                start_row = row
                end_row = row
                while end_row < sheet.max_row and str(sheet.cell(end_row + 1, 1).value).split('_')[0] == current_value:
                    end_row += 1
                if start_row != end_row:
                    for col in range(3, sheet.max_column + 1):
                        column_sum = sum(sheet.cell(r, col).value for r in range(start_row, end_row + 1))
                        average = column_sum / (end_row - start_row + 1)
                        sheet.cell(end_row, 11, value=average)
                        print(f"Average for {current_value}: {average}")

        # Save the changes
        workbook.save(excel_save_path)
        print("Average calculations and updates have been completed.")

        # Print the processed strings to the screen
        processed_strings_output = "\n".join(processed_strings)
        print(f"Processed strings:\n{processed_strings_output}")

        # Open the Excel file again
        workbook = openpyxl.load_workbook(excel_save_path)
        sheet = workbook.active

        # Print all values in column C for each string and their averages
        for string in processed_strings:
            c_values = []
            for row in range(2, sheet.max_row + 1):
                if str(sheet.cell(row, 1).value).split('_')[0] == string:
                    c_values.append(sheet.cell(row, 3).value)
            print(f"Values for {string}: {c_values}")
            print(f"Average for {string}: {sum(c_values) / len(c_values)}")

        # Create a new sheet for metadata
        meta_sheet = workbook.create_sheet("Meta Data")
        for i, line in enumerate(meta_data):
            meta_sheet.cell(row=i+1, column=1, value=line)

        # Create a new sheet for processed data with headers
        new_sheet = workbook.create_sheet("Processed Data")
        headers = ["label", "power", "rel mo", "abs mo", "temp set", "temp rep", "status", "date/time"]
        new_sheet.append(headers)

        for line in buffer.getvalue().split('\n'):
            parts = line.split(":")
            new_sheet.append(parts)

        # Save the changes
        workbook.save(excel_save_path)
        print("Print output and metadata have been saved to Excel.")

        # Reopen the Excel file
        workbook = openpyxl.load_workbook(excel_save_path)
        sheet1 = workbook["Sheet1"]

        # Insert the specified strings into the first row of "Sheet1"
        headers = ["label", "power", "rel mo", "abs mo", "temp set", "temp rep", "status", "date/time"]
        for i, header in enumerate(headers, start=1):
            sheet1.cell(row=1, column=i, value=header)

        # Save the changes
        workbook.save(excel_save_path)
        print("Strings have been inserted into the first row of Sheet1 in the Excel file.")

        # Reopen the Excel file
        workbook = openpyxl.load_workbook(excel_save_path)

        # Access the "Processed Data" sheet
        processed_data_sheet = workbook["Processed Data"]

        # Remove leading spaces from column B
        for row in processed_data_sheet.iter_rows(min_row=2, min_col=2, max_row=processed_data_sheet.max_row, max_col=2):
            for cell in row:
                cell.value = str(cell.value).lstrip()

        # Save the changes
        workbook.save(excel_save_path)
        print("Leading spaces have been removed from column B in the Processed Data sheet.")


    except pd.errors.ParserError as e:
        print(f"Error occurred while reading the file {file_name}: {e}")





